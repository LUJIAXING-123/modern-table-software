import wx
import wx.grid as gridlib
import json
from typing import List, Dict, Any
from openpyxl import load_workbook

class WxTableApp(wx.Frame):
    def __init__(self):
        super().__init__(None, title="现代表格软件", size=(1200, 800))
        
        # 设置主题
        self.SetBackgroundColour(wx.Colour(240, 242, 245))
        
        # 表格数据结构
        self.columns: List[Dict[str, Any]] = []
        self.rows: List[Dict[str, Any]] = []
        
        # 创建主面板
        self.main_panel = wx.Panel(self)
        
        # 创建垂直布局
        self.main_sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 创建工具栏
        self.create_toolbar()
        
        # 创建表格区域
        self.create_table_area()
        
        # 初始化默认表格
        self.add_column()
        self.add_row()
        
        # 更新界面
        self.update_table()
        
        # 设置主面板布局
        self.main_panel.SetSizer(self.main_sizer)
        
        # 居中显示
        self.Center()
    
    def create_toolbar(self):
        """创建工具栏"""
        toolbar = wx.Panel(self.main_panel, style=wx.BORDER_SUNKEN)
        toolbar.SetBackgroundColour(wx.Colour(240, 242, 245))
        
        # 工具栏布局
        toolbar_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        # 按钮样式
        btn_font = wx.Font(12, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL)
        
        # 按钮列表
        buttons = [
            ("添加列", self.on_add_column),
            ("添加行", self.on_add_row),
            ("删除列", self.on_delete_column),
            ("删除行", self.on_delete_row),
            ("保存数据", self.on_save_data),
            ("加载数据", self.on_load_data),
            ("导入XLSX", self.on_import_xlsx),
        ]
        
        # 创建按钮
        for text, callback in buttons:
            button = wx.Button(toolbar, label=text)
            button.SetFont(btn_font)
            button.SetBackgroundColour(wx.Colour(52, 152, 219))
            button.SetForegroundColour(wx.Colour(255, 255, 255))
            button.Bind(wx.EVT_BUTTON, callback)
            toolbar_sizer.Add(button, 0, wx.ALL, 5)
        
        toolbar.SetSizer(toolbar_sizer)
        toolbar.Layout()
        self.main_sizer.Add(toolbar, 0, wx.EXPAND | wx.BOTTOM, 10)
    
    def create_table_area(self):
        """创建表格区域"""
        # 表格容器
        table_panel = wx.Panel(self.main_panel, style=wx.BORDER_SUNKEN)
        table_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        
        # 表格布局
        table_sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 表格标题
        title = wx.StaticText(table_panel, label="表格内容")
        title.SetFont(wx.Font(14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        title.SetForegroundColour(wx.Colour(44, 62, 80))
        table_sizer.Add(title, 0, wx.ALIGN_LEFT | wx.BOTTOM | wx.ALL, 10)
        
        # 创建网格表格
        self.grid = gridlib.Grid(table_panel)
        table_sizer.Add(self.grid, 1, wx.EXPAND | wx.ALL, 10)
        
        table_panel.SetSizer(table_sizer)
        table_panel.Layout()
        self.main_sizer.Add(table_panel, 1, wx.EXPAND | wx.BOTTOM, 10)
        
        # 绑定双击事件
        self.grid.Bind(gridlib.EVT_GRID_CELL_LEFT_DCLICK, self.on_cell_double_click)
        # 绑定列标题双击事件用于修改列名
        self.grid.Bind(gridlib.EVT_GRID_LABEL_LEFT_DCLICK, self.on_col_label_double_click)
    

    
    def on_add_column(self, event):
        """添加列按钮事件"""
        self.add_column()
        self.update_table()
    
    def on_add_row(self, event):
        """添加行按钮事件"""
        self.add_row()
        self.update_table()
    
    def on_delete_column(self, event):
        """删除列按钮事件"""
        if len(self.columns) <= 1:
            wx.MessageBox("至少需要保留一列", "警告", wx.OK | wx.ICON_WARNING)
            return
        
        # 获取选中的列
        selected_cols = self.grid.GetSelectedCols()
        if not selected_cols:
            wx.MessageBox("请选择要删除的列", "警告", wx.OK | wx.ICON_WARNING)
            return
        
        # 删除选中的列
        col_index = selected_cols[0]
        del_col = self.columns.pop(col_index)
        
        # 删除所有行中的该列数据
        for row in self.rows:
            if del_col["id"] in row:
                del row[del_col["id"]]
        
        # 更新表格
        self.update_table()
    
    def on_delete_row(self, event):
        """删除行按钮事件"""
        if len(self.rows) <= 1:
            wx.MessageBox("至少需要保留一行", "警告", wx.OK | wx.ICON_WARNING)
            return
        
        # 获取选中的行
        selected_rows = self.grid.GetSelectedRows()
        if not selected_rows:
            wx.MessageBox("请选择要删除的行", "警告", wx.OK | wx.ICON_WARNING)
            return
        
        # 删除选中的行，从后往前删除避免索引问题
        selected_rows.sort(reverse=True)
        for row_index in selected_rows:
            self.rows.pop(row_index)
        
        # 更新表格
        self.update_table()
    
    def on_save_data(self, event):
        """保存数据按钮事件"""
        with wx.FileDialog(
            self, "保存数据", wildcard="JSON Files (*.json)|*.json|All Files (*)|*",
            style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT
        ) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            
            path = fileDialog.GetPath()
            data = {
                "columns": self.columns,
                "rows": self.rows
            }
            
            try:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                wx.MessageBox("数据已保存", "成功", wx.OK | wx.ICON_INFORMATION)
            except Exception as e:
                wx.MessageBox(f"保存失败: {str(e)}", "错误", wx.OK | wx.ICON_ERROR)
    
    def on_load_data(self, event):
        """加载数据按钮事件"""
        with wx.FileDialog(
            self, "加载数据", wildcard="JSON Files (*.json)|*.json|All Files (*)|*",
            style=wx.FD_OPEN
        ) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            
            path = fileDialog.GetPath()
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                
                if "columns" in data and "rows" in data:
                    self.columns = data["columns"]
                    self.rows = data["rows"]
                    self.update_table()
                    wx.MessageBox("数据已加载", "成功", wx.OK | wx.ICON_INFORMATION)
                else:
                    wx.MessageBox("无效的数据格式", "错误", wx.OK | wx.ICON_ERROR)
            except Exception as e:
                wx.MessageBox(f"加载失败: {str(e)}", "错误", wx.OK | wx.ICON_ERROR)
    
    def on_import_xlsx(self, event):
        """导入XLSX文件按钮事件"""
        with wx.FileDialog(
            self, "导入XLSX文件", wildcard="Excel Files (*.xlsx)|*.xlsx|All Files (*)|*",
            style=wx.FD_OPEN
        ) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            
            path = fileDialog.GetPath()
            try:
                # 加载工作簿
                workbook = load_workbook(path)
                sheet = workbook.active
                
                # 清空现有数据
                self.columns.clear()
                self.rows.clear()
                
                # 读取列标题（第一行）
                columns_data = []
                # 从第一行获取所有列标题
                for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                    columns_data = list(row)
                    break
                
                # 创建列定义
                for i, col_name in enumerate(columns_data):
                    col_id = f"col_{i}"
                    self.columns.append({
                        "id": col_id,
                        "name": str(col_name) if col_name else f"列{i+1}"
                    })
                
                # 读取数据行
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=0):
                    row_id = f"row_{row_idx}"
                    row_data = {"id": row_id}
                    
                    # 初始化行数据
                    for i, col in enumerate(self.columns):
                        if i < len(row):
                            value = str(row[i]) if row[i] is not None else ""
                        else:
                            value = ""
                        row_data[col["id"]] = value
                    
                    self.rows.append(row_data)
                
                # 更新界面
                self.update_table()
                wx.MessageBox(f"已导入 {len(self.rows)} 行数据", "成功", wx.OK | wx.ICON_INFORMATION)
            except Exception as e:
                wx.MessageBox(f"导入失败: {str(e)}", "错误", wx.OK | wx.ICON_ERROR)
    
    def add_column(self):
        """添加列"""
        col_id = f"col_{len(self.columns)}"
        col_name = f"列{len(self.columns) + 1}"
        
        # 添加列定义
        self.columns.append({
            "id": col_id,
            "name": col_name
        })
        
        # 更新所有行数据
        for row in self.rows:
            row[col_id] = ""
    
    def add_row(self):
        """添加行"""
        row_id = f"row_{len(self.rows)}"
        row_data = {"id": row_id}
        
        # 初始化行数据
        for col in self.columns:
            row_data[col["id"]] = ""
        
        self.rows.append(row_data)
    
    def update_table(self):
        """更新表格显示"""
        # 设置表格大小
        num_rows = len(self.rows)
        num_cols = len(self.columns)
        
        # 检查网格是否已经创建，如果创建了则直接更新内容，否则创建新网格
        try:
            # 先获取当前行数和列数
            current_rows = self.grid.GetNumberRows()
            current_cols = self.grid.GetNumberCols()
            
            if current_rows > 0 or current_cols > 0:
                # 更新行数量
                if num_rows > current_rows:
                    self.grid.AppendRows(num_rows - current_rows)
                elif num_rows < current_rows:
                    self.grid.DeleteRows(num_rows, current_rows - num_rows)
                
                # 更新列数量
                if num_cols > current_cols:
                    self.grid.AppendCols(num_cols - current_cols)
                elif num_cols < current_cols:
                    self.grid.DeleteCols(num_cols, current_cols - num_cols)
                
                # 更新所有列的标题和宽度（无论列数是否变化）
                for i, col in enumerate(self.columns):
                    self.grid.SetColLabelValue(i, col["name"])
                    self.grid.SetColSize(i, 150)
            else:
                # 创建新网格
                self.grid.CreateGrid(num_rows, num_cols)
                # 设置列标题
                for i, col in enumerate(self.columns):
                    self.grid.SetColLabelValue(i, col["name"])
                    self.grid.SetColSize(i, 150)
            
            # 更新所有单元格内容
            for row_idx in range(num_rows):
                for col_idx in range(num_cols):
                    self.grid.SetCellValue(row_idx, col_idx, self.rows[row_idx].get(self.columns[col_idx]["id"], ""))
        except Exception as e:
            # 如果发生错误，尝试重新创建网格
            try:
                self.grid.ClearGrid()
                self.grid.DeleteGrid()
            except:
                pass
            # 创建新网格
            self.grid.CreateGrid(num_rows, num_cols)
            # 设置列标题
            for i, col in enumerate(self.columns):
                self.grid.SetColLabelValue(i, col["name"])
                self.grid.SetColSize(i, 150)
            # 设置行数据
            for row_idx, row in enumerate(self.rows):
                for col_idx, col in enumerate(self.columns):
                    self.grid.SetCellValue(row_idx, col_idx, row.get(col["id"], ""))
    

    
    def on_cell_double_click(self, event):
        """处理单元格双击事件"""
        row_idx = event.GetRow()
        col_idx = event.GetCol()
        
        if row_idx >= len(self.rows) or col_idx >= len(self.columns):
            return
        
        # 获取当前值
        current_value = self.rows[row_idx][self.columns[col_idx]["id"]]
        
        # 创建编辑对话框
        self.create_edit_dialog(row_idx, col_idx, current_value)
    
    def on_col_label_double_click(self, event):
        """处理列标题双击事件，用于修改列名"""
        # 检查是否是列标题双击
        row = event.GetRow()
        col = event.GetCol()
        
        # 行号为-1表示是列标题
        if row == -1 and col < len(self.columns):
            # 获取当前列名
            current_col_name = self.columns[col]["name"]
            
            # 创建修改列名对话框
            self.create_edit_column_name_dialog(col, current_col_name)
    
    def create_edit_dialog(self, row_idx, col_idx, current_value):
        """创建编辑对话框"""
        col = self.columns[col_idx]
        
        # 创建对话框
        dialog = wx.Dialog(self, title=f"编辑 {col['name']}", size=(400, 200))
        
        # 对话框布局
        dialog_sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 标题
        title = wx.StaticText(dialog, label=f"编辑 {col['name']}")
        title.SetFont(wx.Font(14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        title.SetForegroundColour(wx.Colour(44, 62, 80))
        dialog_sizer.Add(title, 0, wx.ALIGN_CENTER | wx.BOTTOM | wx.ALL, 10)
        
        # 输入控件 - 只使用文本输入
        text_ctrl = wx.TextCtrl(dialog, value=str(current_value))
        dialog_sizer.Add(text_ctrl, 0, wx.EXPAND | wx.ALL, 10)
        
        # 按钮布局
        button_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        # 取消按钮
        cancel_btn = wx.Button(dialog, label="取消")
        cancel_btn.Bind(wx.EVT_BUTTON, lambda e: dialog.Close())
        button_sizer.Add(cancel_btn, 0, wx.ALIGN_CENTER | wx.RIGHT, 10)
        
        # 保存按钮
        save_btn = wx.Button(dialog, label="保存")
        
        def on_save():
            new_value = text_ctrl.GetValue()
            self.rows[row_idx][col["id"]] = new_value
            self.update_table()
            dialog.Close()
        
        save_btn.Bind(wx.EVT_BUTTON, lambda e: on_save())
        button_sizer.Add(save_btn, 0, wx.ALIGN_CENTER)
        
        dialog_sizer.Add(button_sizer, 0, wx.ALIGN_CENTER | wx.TOP | wx.BOTTOM, 10)
        
        # 设置对话框布局
        dialog.SetSizer(dialog_sizer)
        dialog.Layout()
        
        # 显示对话框
        dialog.ShowModal()
        dialog.Destroy()
    
    def create_edit_column_name_dialog(self, col_idx, current_col_name):
        """创建修改列名对话框"""
        # 创建对话框
        dialog = wx.Dialog(self, title="修改列名", size=(400, 200))
        
        # 对话框布局
        dialog_sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 标题
        title = wx.StaticText(dialog, label="修改列名")
        title.SetFont(wx.Font(14, wx.FONTFAMILY_SWISS, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD))
        title.SetForegroundColour(wx.Colour(44, 62, 80))
        dialog_sizer.Add(title, 0, wx.ALIGN_CENTER | wx.BOTTOM | wx.ALL, 10)
        
        # 列名输入控件
        text_ctrl = wx.TextCtrl(dialog, value=str(current_col_name))
        text_ctrl.SetFocus()  # 设置焦点，方便用户直接输入
        text_ctrl.SelectAll()  # 选中现有文本，方便替换
        dialog_sizer.Add(text_ctrl, 0, wx.EXPAND | wx.ALL, 10)
        
        # 按钮布局
        button_sizer = wx.BoxSizer(wx.HORIZONTAL)
        
        # 取消按钮
        cancel_btn = wx.Button(dialog, label="取消")
        cancel_btn.Bind(wx.EVT_BUTTON, lambda e: dialog.Close())
        button_sizer.Add(cancel_btn, 0, wx.ALIGN_CENTER | wx.RIGHT, 10)
        
        # 保存按钮
        save_btn = wx.Button(dialog, label="保存")
        
        def on_save():
            new_col_name = text_ctrl.GetValue().strip()
            if not new_col_name:
                wx.MessageBox("列名不能为空", "警告", wx.OK | wx.ICON_WARNING)
                return
            
            # 更新列名
            self.columns[col_idx]["name"] = new_col_name
            # 更新表格显示
            self.update_table()
            dialog.Close()
        
        save_btn.Bind(wx.EVT_BUTTON, lambda e: on_save())
        button_sizer.Add(save_btn, 0, wx.ALIGN_CENTER)
        
        dialog_sizer.Add(button_sizer, 0, wx.ALIGN_CENTER | wx.TOP | wx.BOTTOM, 10)
        
        # 设置对话框布局
        dialog.SetSizer(dialog_sizer)
        dialog.Layout()
        
        # 显示对话框
        dialog.ShowModal()
        dialog.Destroy()

if __name__ == "__main__":
    app = wx.App(False)
    frame = WxTableApp()
    frame.Show()
    app.MainLoop()